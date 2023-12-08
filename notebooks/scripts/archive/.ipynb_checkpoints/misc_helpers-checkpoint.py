# IMPORTS
import os
import shutil
import time
import numpy as np
from scripts import audio_control
from scripts import ucb1_algorithm as ucb1

import sys
from termcolor import colored, cprint
# Termcolor guide: https://pypi.org/project/termcolor/

import random

from openpyxl import Workbook
from openpyxl import load_workbook

# https://www.blog.pythonlibrary.org/2021/07/27/creating-spreadsheets-with-openpyxl-and-python/

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# INITIALIZATIONS
# None 

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# STAND-ALONE FUNCTIONS

def get_user_ID(parent_dir, num_of_states):
	''' returns an string of the double digit user ID '''
	
	user_ID_str = "-1" # <---"enter your number here"
	
	while True:
		try:
			user_ID_str = str(input(f"\nPlease enter your double digiet user ID number (ex: 00 ) then hit 'enter'\n"))

		except ValueError:
			print("Invalid user ID...\n")
			continue

		if user_ID_str != "-1" and len(user_ID_str) == 2:
			
			directory = "user_data/user_" + user_ID_str + "/arrays"

			# Path (add parent directory with new folder)
			path = os.path.join(parent_dir, directory)

			# Create the directory
			try:
				os.makedirs(path, exist_ok = True)
				# print("\nDirectory '%s' created successfully" % directory)
			except OSError as error:
				# print("\nDirectory '%s' can not be created" % directory)
				pass
		

			for state_idx in range(num_of_states):
				shutil.copyfile("arrays/pilotset_st" + str(state_idx) + ".npy", "user_data/user_" + user_ID_str + "/arrays/pilotset_st" + str(state_idx) + ".npy")

			# Coloured print statement to direct user to next cell
			cprint("\n\n\n------------------------------------------------------------------------", "light_yellow", attrs=["bold"])
			cprint("------------------------------------------------------------------------\n", "light_yellow", attrs=["bold"])
			cprint(f"Great job! You are user: {user_ID_str}\n", "black", "on_yellow", attrs=["bold"])
			cprint(f"Click on the next cell below and hit 'shift + enter' to continue\n", "black", "on_yellow", attrs=["bold"])
			cprint("------------------------------------------------------------------------", "light_yellow", attrs=["bold"])
			cprint("------------------------------------------------------------------------\n\n\n", "light_yellow", attrs=["bold"])
			
			break

		else:
			print("Invalid user ID...\n")
	
	time.sleep(1)
	
	return user_ID_str


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


def get_user_accuracy(sound_obj_array, lib_str, sect_str, user_ID_str, num_of_states, states_array, state_descriptions, param_disc, load_file="pilotset", seed=55, mixer_volume=0.70):
	''' get the accuracy of the user given an existing set of states, sound library and Q-table for each state. Saves responses into an excel spreadsheet.'''
	
	random.seed(seed)

	
	rand_state_idx_list = [*range(0, num_of_states, 1)]
	random.shuffle(rand_state_idx_list)
	# print("Suffled rand_state_idx_list:", rand_state_idx_list)

	uncertainty_array = np.zeros_like(sound_obj_array)

	
	# Enter the data in spreadsheet format
	workbook_path = "user_data/response_book.xlsx"
	response_book = load_workbook(workbook_path)
	
	try: # Try to open existing sheet
		response_sheet = response_book.get_sheet_by_name("user_" + user_ID_str)
	except KeyError:  # If ot doesn't exist. create it
		response_sheet = response_book.create_sheet(title="user_" + user_ID_str)
	
	response_sheet["B1"] = "Actual State"
	response_sheet["C1"] = "Probed State"
	response_sheet["D1"] = "Probed Confidence"
	
	response_sheet["A2"] = "sect1 libA"
	response_sheet["A6"] = "sect1 libB"
	response_sheet["A10"] = "sect3 libA"
	response_sheet["A14"] = "sect3 libB"

		
	# Creating a text file with the command function "w" (this is more or less just a backup)
	textfile = open("user_data/user_" + user_ID_str + "/" + sect_str + "_" + lib_str + "_responses.txt", "w")
	textfile.flush()

	
	for state_idx in rand_state_idx_list:  # To run them in order, replace with: for ... in range(num_of_states)
		current_state_index = state_idx

		# At each index of the state array, add the state object (each object has a unique index, description, and Q-table)
		states_array[state_idx] = ucb1.robot_state(state_idx, state_descriptions[state_idx], param_disc, load_file, user_ID_str)

		# Select the highest valued action in that states Q-Table - assign to params 1,2,3
		param_1_idx, param_2_idx, param_3_idx = states_array[current_state_index].action_selection(uncertainty_array)

		# Now lets play this action for the user and get their reponse 
		probed_state_index, probed_confidence = sound_obj_array[param_1_idx, param_2_idx, param_3_idx].probe(state_descriptions, mixer_volume)


		# Enter the data in spreadsheet format
		if sect_str == "sect1":
			if lib_str == "libA":
				if state_idx == 0:
					response_sheet["B2"] = current_state_index
					response_sheet["C2"] = probed_state_index
					response_sheet["D2"] = probed_confidence

				if state_idx == 1:
					response_sheet["B3"] = current_state_index
					response_sheet["C3"] = probed_state_index
					response_sheet["D3"] = probed_confidence		

				if state_idx == 2:
					response_sheet["B4"] = current_state_index
					response_sheet["C4"] = probed_state_index
					response_sheet["D4"] = probed_confidence
		
			if lib_str == "libB":
				if state_idx == 0:
					response_sheet["B6"] = current_state_index
					response_sheet["C6"] = probed_state_index
					response_sheet["D6"] = probed_confidence

				if state_idx == 1:
					response_sheet["B7"] = current_state_index
					response_sheet["C7"] = probed_state_index
					response_sheet["D7"] = probed_confidence
				if state_idx == 2:
					response_sheet["B8"] = current_state_index
					response_sheet["C8"] = probed_state_index
					response_sheet["D8"] = probed_confidence
		
		
		if sect_str == "sect3":
			if lib_str == "libA":
				if state_idx == 0:
					response_sheet["B10"] = current_state_index
					response_sheet["C10"] = probed_state_index
					response_sheet["D10"] = probed_confidence

				if state_idx == 1:
					response_sheet["B11"] = current_state_index
					response_sheet["C11"] = probed_state_index
					response_sheet["D11"] = probed_confidence

				if state_idx == 2:
					response_sheet["B12"] = current_state_index
					response_sheet["C12"] = probed_state_index
					response_sheet["D12"] = probed_confidence
		
			if lib_str == "libB":
				if state_idx == 0:
					response_sheet["B14"] = current_state_index
					response_sheet["C14"] = probed_state_index
					response_sheet["D14"] = probed_confidence

				if state_idx == 1:
					response_sheet["B15"] = current_state_index
					response_sheet["C15"] = probed_state_index
					response_sheet["D15"] = probed_confidence

				if state_idx == 2:
					response_sheet["B16"] = current_state_index
					response_sheet["C16"] = probed_state_index
					response_sheet["D16"] = probed_confidence
		
		
		textfile.write(f"{current_state_index:02} - current_state_index \n")
		textfile.write(f"{probed_state_index:02} - probed_state_index \n")
		textfile.write(f"{probed_confidence:02} - probed_confidence \n\n")
		textfile.flush()
	
	
	response_book.save(workbook_path)

	textfile.close()

	# Coloured print statement to direct user to next cell
	cprint("\n\n\n------------------------------------------------------------------------", "light_yellow", attrs=["bold"])
	cprint("------------------------------------------------------------------------\n", "light_yellow", attrs=["bold"])
	cprint(f"Great job!", "black", "on_yellow", attrs=["bold"])
	cprint(f"Click on the next cell below and hit 'shift + enter' to continue\n", "black", "on_yellow", attrs=["bold"])
	cprint("------------------------------------------------------------------------", "light_yellow", attrs=["bold"])
	cprint("------------------------------------------------------------------------\n\n\n", "light_yellow", attrs=["bold"])
